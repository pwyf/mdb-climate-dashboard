"""
update-data.py
==============
Run this script whenever the Excel dataset is updated.

Usage:
    /c/Users/EllaRemande-Guyard/anaconda3/python.exe update-data.py

It will:
  1. Find the Excel file in the parent folder (any .xlsx matching 'mdb-climate')
  2. Extract all records from the '1. Enriched Dataset' sheet
  3. Inject the new JSON data into index.html
  4. Print a summary of what was loaded

After running, commit and push index.html to update the live dashboard.
"""

import json
import os
import glob
import sys
from datetime import datetime

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR   = os.path.dirname(SCRIPT_DIR)
INDEX_HTML   = os.path.join(SCRIPT_DIR, 'index.html')
SHEET_NAME   = '1. Enriched Dataset'

# ── Find Excel file ───────────────────────────────────────────────────────────
candidates = glob.glob(os.path.join(PARENT_DIR, 'mdb-climate*.xlsx'))
if not candidates:
    # Fallback: look in the same folder as this script
    candidates = glob.glob(os.path.join(SCRIPT_DIR, 'mdb-climate*.xlsx'))
if not candidates:
    print("ERROR: No Excel file matching 'mdb-climate*.xlsx' found in:")
    print(f"  {PARENT_DIR}")
    print(f"  {SCRIPT_DIR}")
    print("\nPlace the updated Excel file in the same folder as this script or one level up.")
    sys.exit(1)

# If multiple matches, use the most recently modified one
excel_path = max(candidates, key=os.path.getmtime)
print(f"Using Excel file: {os.path.basename(excel_path)}")

# ── Load Excel ────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

print(f"Loading sheet '{SHEET_NAME}'...")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
if SHEET_NAME not in wb.sheetnames:
    print(f"ERROR: Sheet '{SHEET_NAME}' not found.")
    print(f"Available sheets: {wb.sheetnames}")
    sys.exit(1)

ws = wb[SHEET_NAME]

# ── Column mapping ────────────────────────────────────────────────────────────
# Maps Excel column header -> internal field name used by the dashboard JS
COL_MAP = {
    'MDB':                         'mdb',
    'Sovereign / non-sovereign ':  'sovereign',
    'Project name':                'project_name',
    'Project ID':                  'project_id',
    'Approval / signature date':   'approval_date',
    'Approval / reporting year':   'year',
    'Country':                     'country',
    'Investment instrument':       'instrument',
    'Sector 1':                    'sector',
    'Type':                        'type',
    'Total commitment ($ million)':'total_commitment',
    'Climate finance ($ million)': 'climate_finance',
    'Mitigation ($ million)':      'mitigation',
    'Adaptation ($ million)':      'adaptation',
    'Dual-use ($ million)':        'dual_use',
    'Climate finance (%)':         'climate_finance_pct',
    'Project URL':                 'url',
}

# ── Read headers ──────────────────────────────────────────────────────────────
rows = ws.iter_rows(values_only=True)
raw_headers = list(next(rows))

# Build index: field_name -> column index (skip unmapped headers)
col_index = {}
for i, h in enumerate(raw_headers):
    if h in COL_MAP:
        col_index[COL_MAP[h]] = i

missing = [k for k in COL_MAP.values() if k not in col_index]
if missing:
    print(f"WARNING: These expected fields were not found in the sheet: {missing}")
    print("The dashboard may be missing some data. Check the column headers in the Excel file.")

# ── Extract records ───────────────────────────────────────────────────────────
NUMERIC_FIELDS = {'total_commitment', 'climate_finance', 'mitigation',
                  'adaptation', 'dual_use', 'climate_finance_pct', 'year'}

records = []
skipped = 0

for row in rows:
    # Skip completely empty rows
    if all(v is None for v in row):
        skipped += 1
        continue

    rec = {}
    for field, idx in col_index.items():
        val = row[idx] if idx < len(row) else None

        if field == 'approval_date':
            # Format as DD/MM/YYYY string
            if isinstance(val, datetime):
                val = val.strftime('%d/%m/%Y')
            elif val is not None:
                val = str(val).strip() or None

        elif field == 'sovereign':
            # Normalise to "Sovereign" / "Nonsovereign"
            if val is not None:
                s = str(val).strip().lower()
                if 'non' in s:
                    val = 'Nonsovereign'
                else:
                    val = 'Sovereign'

        elif field == 'country':
            if val is not None:
                COUNTRY_ALIASES = {'ACP': 'Regional', 'Viet nam': 'Viet Nam', 'Viet Nam)': 'Viet Nam'}
                parts = [COUNTRY_ALIASES.get(p.strip(), p.strip()) for p in str(val).split(';') if p.strip()]
                parts = list(dict.fromkeys(parts))  # deduplicate, preserve order
                val = '; '.join(parts) or None

        elif field in NUMERIC_FIELDS:
            if val is not None:
                try:
                    val = float(val)
                    if field == 'year':
                        val = int(val)
                except (ValueError, TypeError):
                    val = None

        else:
            if val is not None:
                val = str(val).strip() or None

        rec[field] = val

    # Skip rows with no project name (likely header repetitions or footers)
    if not rec.get('project_name'):
        skipped += 1
        continue

    records.append(rec)

wb.close()
print(f"Extracted {len(records):,} records ({skipped} rows skipped)")

# ── Inject into index.html ────────────────────────────────────────────────────
print("Reading index.html...")
with open(INDEX_HTML, 'r', encoding='utf-8') as f:
    html = f.read()

MARKER_START = '<script>const MDB_DATA = '
MARKER_END   = ';</script>'

si = html.index(MARKER_START)
ei = html.index(MARKER_END, si)

new_json = json.dumps(records, ensure_ascii=False, separators=(',', ':'))
html = html[:si] + MARKER_START + new_json + MARKER_END + html[ei + len(MARKER_END):]

print("Writing index.html...")
with open(INDEX_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

size_mb = round(len(html) / 1024 / 1024, 2)

# ── Summary ───────────────────────────────────────────────────────────────────
years  = sorted({r['year'] for r in records if r.get('year')})
mdbs   = sorted({r['mdb']  for r in records if r.get('mdb')})
total_cf = sum(r['climate_finance'] for r in records if r.get('climate_finance'))

print()
print("=" * 50)
print(f"  Records loaded : {len(records):,}")
print(f"  Years covered  : {years[0]}–{years[-1]}" if years else "  Years: unknown")
print(f"  MDBs           : {len(mdbs)}")
print(f"  Total CF       : ${total_cf:,.1f}M")
print(f"  index.html size: {size_mb} MB")
print("=" * 50)
print()
print("Done! Next steps:")
print("  git add index.html")
print('  git commit -m "Update data to latest Excel"')
print("  git push")
